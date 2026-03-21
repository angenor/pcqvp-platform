"""Import des structures de reference des comptes administratifs depuis Excel.

Usage:
    python -m app.services.seed_templates
"""

import asyncio
from pathlib import Path

import openpyxl
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import async_session
from app.models.account_template import (
    AccountTemplate,
    AccountTemplateColumn,
    AccountTemplateLine,
    ColumnDataType,
    SectionType,
    TemplateType,
)

DATA_DIR = Path(__file__).parent.parent / "data" / "reference"
TEMPLATE_FILE = DATA_DIR / "Template_Tableaux_de_Compte_Administratif.xlsx"

RECETTES_COLUMNS = [
    ("budget_primitif", "Budget primitif", ColumnDataType.number, False, None),
    ("budget_additionnel", "Budget additionnel", ColumnDataType.number, False, None),
    ("modifications", "Modifications +/-", ColumnDataType.number, False, None),
    (
        "previsions_definitives",
        "Previsions definitives",
        ColumnDataType.number,
        True,
        "budget_primitif + budget_additionnel + modifications",
    ),
    ("or_admis", "OR admis", ColumnDataType.number, False, None),
    ("recouvrement", "Recouvrement", ColumnDataType.number, False, None),
    (
        "reste_a_recouvrer",
        "Reste a recouvrer",
        ColumnDataType.number,
        True,
        "or_admis - recouvrement",
    ),
    (
        "taux_execution",
        "Taux d'execution",
        ColumnDataType.percentage,
        True,
        "or_admis / previsions_definitives",
    ),
]

DEPENSES_COLUMNS = [
    ("budget_primitif", "Budget primitif", ColumnDataType.number, False, None),
    ("budget_additionnel", "Budget additionnel", ColumnDataType.number, False, None),
    ("modifications", "Modifications +/-", ColumnDataType.number, False, None),
    (
        "previsions_definitives",
        "Previsions definitives",
        ColumnDataType.number,
        True,
        "budget_primitif + budget_additionnel + modifications",
    ),
    ("engagement", "Engagement", ColumnDataType.number, False, None),
    ("mandat_admis", "Mandat admis", ColumnDataType.number, False, None),
    ("paiement", "Paiement", ColumnDataType.number, False, None),
    (
        "reste_a_payer",
        "Reste a payer",
        ColumnDataType.number,
        True,
        "mandat_admis - paiement",
    ),
    (
        "taux_execution",
        "Taux d'execution",
        ColumnDataType.percentage,
        True,
        "mandat_admis / previsions_definitives",
    ),
]


def _parse_sheet_lines(ws) -> list[dict]:
    """Parse les lignes de comptes d'une feuille Excel."""
    lines = []
    current_section = None
    sort_order = 0
    seen_codes: set[str] = set()

    for row_idx in range(1, ws.max_row + 1):
        col_a = ws.cell(row=row_idx, column=1).value
        col_b = ws.cell(row=row_idx, column=2).value
        col_c = ws.cell(row=row_idx, column=3).value
        col_d = ws.cell(row=row_idx, column=4).value
        col_e = ws.cell(row=row_idx, column=5).value

        # Detect section headers (can be in col A, B, or E)
        row_text = " ".join(
            str(v or "").upper()
            for v in [col_a, col_b, col_e]
        )
        if "FONCTIONNEMENT" in row_text and "TOTAL" not in row_text:
            # Check if col_b is text (header) not a numeric code
            b_str = str(col_b or "").strip()
            if not b_str or not b_str[0].isdigit():
                current_section = SectionType.fonctionnement
                continue
        if "INVESTISSEMENT" in row_text and "TOTAL" not in row_text:
            b_str = str(col_b or "").strip()
            if not b_str or not b_str[0].isdigit():
                current_section = SectionType.investissement
                continue

        if current_section is None:
            continue

        # Determine level and code
        code = None
        level = None
        if col_b is not None and str(col_b).strip():
            b_str = str(col_b).strip()
            # Skip non-numeric values (column headers like "COMPTE")
            if b_str[0].isdigit():
                code = b_str
                level = 1
        if code is None and col_c is not None and str(col_c).strip():
            c_str = str(col_c).strip()
            if c_str[0].isdigit():
                code = c_str
                level = 2
        if code is None and col_d is not None and str(col_d).strip():
            d_str = str(col_d).strip()
            if d_str[0].isdigit():
                code = d_str
                level = 3

        if code is None or level is None:
            continue

        # Skip total/header rows
        intitule = str(col_e or "").strip()
        if not intitule or "TOTAL" in intitule.upper():
            continue
        if "SOUS-TOTAL" in intitule.upper():
            continue

        # Skip duplicates (totals sometimes repeat codes)
        if code in seen_codes:
            continue
        seen_codes.add(code)

        # Determine parent_code
        parent_code = None
        if level == 2:
            parent_code = code[:2]
        elif level == 3:
            parent_code = code[:3]

        sort_order += 1
        lines.append(
            {
                "compte_code": code,
                "intitule": intitule,
                "level": level,
                "parent_code": parent_code,
                "section": current_section,
                "sort_order": sort_order,
            }
        )

    return lines


async def _upsert_template(
    db: AsyncSession,
    name: str,
    template_type: TemplateType,
    lines_data: list[dict],
    columns_def: list[tuple],
) -> AccountTemplate:
    """Cree ou met a jour un template avec ses lignes et colonnes."""
    result = await db.execute(
        select(AccountTemplate).where(
            AccountTemplate.name == name,
            AccountTemplate.type == template_type,
        )
    )
    template = result.scalar_one_or_none()

    if template is None:
        template = AccountTemplate(name=name, type=template_type)
        db.add(template)
        await db.flush()
    else:
        # Clear existing lines and columns for re-import
        await db.execute(
            select(AccountTemplateLine)
            .where(AccountTemplateLine.template_id == template.id)
        )
        for line in list(template.lines):
            await db.delete(line)
        for col in list(template.columns):
            await db.delete(col)
        await db.flush()

    # Create lines
    for line_data in lines_data:
        line = AccountTemplateLine(
            template_id=template.id,
            compte_code=line_data["compte_code"],
            intitule=line_data["intitule"],
            level=line_data["level"],
            parent_code=line_data["parent_code"],
            section=line_data["section"],
            sort_order=line_data["sort_order"],
        )
        db.add(line)

    # Create columns
    for idx, (code, col_name, data_type, is_computed, formula) in enumerate(
        columns_def, start=1
    ):
        col = AccountTemplateColumn(
            template_id=template.id,
            name=col_name,
            code=code,
            data_type=data_type,
            is_computed=is_computed,
            formula=formula,
            sort_order=idx,
        )
        db.add(col)

    await db.commit()
    await db.refresh(template)
    return template


async def seed_templates(db: AsyncSession | None = None) -> dict:
    """Importe les templates de reference depuis le fichier Excel."""
    close_session = False
    if db is None:
        session_ctx = async_session()
        db = await session_ctx.__aenter__()
        close_session = True

    try:
        wb = openpyxl.load_workbook(TEMPLATE_FILE, data_only=True)

        # Parse RECETTES
        ws_recettes = wb["RECETTES"]
        recettes_lines = _parse_sheet_lines(ws_recettes)

        # Parse DEPENSES (Programme I as reference structure)
        ws_depenses = wb["DEPENSES PROGRAMME I"]
        depenses_lines = _parse_sheet_lines(ws_depenses)

        wb.close()

        # Upsert templates
        t_recettes = await _upsert_template(
            db, "Recettes", TemplateType.recette, recettes_lines, RECETTES_COLUMNS
        )
        t_depenses = await _upsert_template(
            db, "Depenses", TemplateType.depense, depenses_lines, DEPENSES_COLUMNS
        )

        result = {
            "recettes": {
                "id": str(t_recettes.id),
                "lines": len(recettes_lines),
                "columns": len(RECETTES_COLUMNS),
            },
            "depenses": {
                "id": str(t_depenses.id),
                "lines": len(depenses_lines),
                "columns": len(DEPENSES_COLUMNS),
            },
        }
        print(f"Recettes: {result['recettes']['lines']} lignes, "
              f"{result['recettes']['columns']} colonnes")
        print(f"Depenses: {result['depenses']['lines']} lignes, "
              f"{result['depenses']['columns']} colonnes")
        return result
    finally:
        if close_session:
            await db.close()


ANDRAFIABE_FILE = DATA_DIR / "COMPTE_ADMINISTRATIF_COMMUNE_ANDRAFIABE_2023.xlsx"

PROGRAMME_SHEETS = {
    "DEP PROGRAM I": "ADMINISTRATION ET COORDINATION",
    "DEP PROGRAM II": "DEVELOPPEMENT ECONOMIQUE ET SOCIAL",
    "DEP PROGRAM III": "SANTE",
}


def parse_andrafiabe_recettes() -> list[dict]:
    """Parse Andrafiabe recette data (code in B, values in D-K)."""
    wb = openpyxl.load_workbook(ANDRAFIABE_FILE, data_only=True)
    ws = wb["RECETTE"]
    rows = []
    for row_idx in range(5, ws.max_row + 1):
        code = ws.cell(row=row_idx, column=2).value
        if code is None:
            continue
        code = str(code).strip()
        if not code or not code[0].isdigit():
            continue
        intitule = str(ws.cell(row=row_idx, column=3).value or "")
        if "total" in intitule.lower():
            continue
        rows.append({
            "code": code,
            "budget_primitif": _num(ws.cell(row=row_idx, column=4).value),
            "budget_additionnel": _num(
                ws.cell(row=row_idx, column=5).value
            ),
            "modifications": _num(ws.cell(row=row_idx, column=6).value),
            "or_admis": _num(ws.cell(row=row_idx, column=8).value),
            "recouvrement": _num(ws.cell(row=row_idx, column=9).value),
        })
    wb.close()
    return rows


def parse_andrafiabe_depenses() -> dict[str, list[dict]]:
    """Parse Andrafiabe depenses data for each programme."""
    wb = openpyxl.load_workbook(ANDRAFIABE_FILE, data_only=True)
    result = {}
    for sheet_name, programme_name in PROGRAMME_SHEETS.items():
        ws = wb[sheet_name]
        rows = []
        for row_idx in range(4, ws.max_row + 1):
            code = ws.cell(row=row_idx, column=2).value
            if code is None:
                continue
            code = str(code).strip()
            if not code or not code[0].isdigit():
                continue
            intitule = str(ws.cell(row=row_idx, column=3).value or "")
            if "total" in intitule.lower():
                continue
            rows.append({
                "code": code,
                "budget_primitif": _num(
                    ws.cell(row=row_idx, column=4).value
                ),
                "budget_additionnel": _num(
                    ws.cell(row=row_idx, column=5).value
                ),
                "modifications": _num(
                    ws.cell(row=row_idx, column=6).value
                ),
                "engagement": _num(ws.cell(row=row_idx, column=8).value),
                "mandat_admis": _num(
                    ws.cell(row=row_idx, column=9).value
                ),
                "paiement": _num(ws.cell(row=row_idx, column=10).value),
            })
        result[programme_name] = rows
    wb.close()
    return result


def _num(value) -> float:
    """Safely convert a cell value to float."""
    if value is None:
        return 0.0
    if isinstance(value, str):
        value = value.strip()
        if not value or value == "#DIV/0!":
            return 0.0
        try:
            return float(value)
        except ValueError:
            return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


async def main():
    print("Importing account templates...")
    result = await seed_templates()
    print(f"Done. {result}")


if __name__ == "__main__":
    asyncio.run(main())
